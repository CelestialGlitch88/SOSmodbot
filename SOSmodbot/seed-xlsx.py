from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), '..', 'logs', 'sos_log.xlsx')

WHITE  = "FFFFFFFF"
HEADER = "FF1A1A2E"
GREY   = "FFF4F6F9"

def hdr(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color="FF444444"))

def widths(sheet, ws):
    for i, w in enumerate(ws, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ── Sheet 1: Strikes ──────────────────────────────────────────────
s1 = wb.active
s1.title = "Strikes"
s1.row_dimensions[1].height = 30
for i, h in enumerate(["Timestamp","Name","Phone Number","Violation Type","Strike #","Day/Night","Content Preview"], 1):
    hdr(s1.cell(1, i), h)
widths(s1, [22,18,16,20,10,10,45])
s1.freeze_panes = "A2"

# ── Sheet 2: Event Log ────────────────────────────────────────────
s2 = wb.create_sheet("Event Log")
s2.row_dimensions[1].height = 30
for i, h in enumerate(["Timestamp","Name","Phone Number","Message Type","Content / Caption","Media Type","Is Forward","Is Violation","Day/Night"], 1):
    hdr(s2.cell(1, i), h)
widths(s2, [22,18,16,14,48,14,12,12,10])
s2.freeze_panes = "A2"

# ── Sheet 3: Admin Roster ─────────────────────────────────────────
s3 = wb.create_sheet("Admin Roster")
s3.row_dimensions[1].height = 30
for i, h in enumerate(["Name","Phone Number (with country code)","Role","Active (YES/NO)","Notes"], 1):
    hdr(s3.cell(1, i), h)

sample = ["Ravi Kumar","919876543210","Top-Level Admin","YES","Building A representative"]
for i, v in enumerate(sample, 1):
    c = s3.cell(2, i)
    c.value = v
    c.font = Font(name="Arial", size=10, italic=True, color="FF888888")
    c.fill = PatternFill("solid", start_color=GREY)

widths(s3, [22,28,20,16,35])

s3["A4"].value = "COLOUR LEGEND"
s3["A4"].font = Font(bold=True, name="Arial", size=10)

for row, (colour, txt) in enumerate([
    ("FFFF4444", "Night violation — 1st strike → DMs sent to all admins"),
    ("FFFFA500", "Day violation — 1st strike (orange)"),
    ("FFFF4444", "Day violation — 2nd+ strike (red)"),
], 5):
    c = s3.cell(row, 1)
    c.value = txt
    c.fill = PatternFill("solid", start_color=colour)
    c.font = Font(name="Arial", size=10, color=WHITE)

s3.freeze_panes = "A2"

wb.save(OUT)
print("XLSX template created at", OUT)
